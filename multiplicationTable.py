#! python3
# multiplicationTable.py takes a number N from the command line
# and creates an N × N multiplication table in an Excel spreadsheet.
# usage: py multiplicationTable.py 6

# Author: Josh Smith

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import sys


class MultTable:

    def __init__(self, number=1):
        wb = Workbook()
        wb_file = f'{number}.xlsx'
        wb.save(wb_file)
        wb = load_workbook(wb_file)
        sheet = wb['Sheet']
        ft = Font(bold=True)
        # row1 listing all integers greater than 1 to the passed integer
        # row1 and colA should be bold
        for i in range(number + 1):
            sheet.cell(row=i + 1, column=1).value = i
            sheet.cell(row=i + 1, column=1).font = ft
            sheet.cell(row=1, column=i + 1).value = i
            sheet.cell(row=1, column=i + 1).font = ft
        # then have products of each row/col in the grid
            for n in range(1, number + 1):
                sheet.cell(row=i + 1, column=n + 1).value = i * n
            sheet.cell(row=1, column=1).value = None
        # save the sheet then finished.
        wb.save(wb_file)


# Run Program when called
if __name__ == "__main__":
    # get valid integer passed via cmd. reject incorrect with error
    if len(sys.argv) < 2:
        sys.exit('Missing required argument. '
                 'Please re-run with integer specified')
    number = (sys.argv[1])
    try:
        number = int(number)
    except ValueError:
        sys.exit(
            'I can only generate a multiplication table with a valid integer.'
            '\nPlease try again.'
        )
    MultTable(number)
    print(f'Done creating Multiplication table for {str(number)}..')


# Code tested and works in cmd.