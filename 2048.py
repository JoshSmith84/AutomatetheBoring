# 2048.py - see how high a score a program can get on 2048
# (https://play2048.co/)
# (by going up, right ,down, and left until game over.) <- Original assignemnt
# I thought I had a better strategy. So why not test it :)
# Will try to see if I can print score.

# Author: Josh Smith

import sys
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import Workbook, load_workbook
from get_integer import get_integer


# Due to some odd behavior, had to add a short wait in between keystrokes.
def wait():
    time.sleep(0.05)


def book_strategy():
    global game
    game.send_keys(Keys.UP)
    wait()
    game.send_keys(Keys.RIGHT)
    wait()
    game.send_keys(Keys.DOWN)
    wait()
    game.send_keys(Keys.LEFT)
    wait()


def josh_strategy():
    global game
    for i in range(10):
        game.send_keys(Keys.DOWN)
        wait()
        game.send_keys(Keys.RIGHT)
        wait()
    game.send_keys(Keys.UP)
    wait()
    game.send_keys(Keys.RIGHT)
    wait()


def update_sheet(sheet_name):
    global old_score
    global wb
    wb = load_workbook(wb_file)
    sheet = wb[sheet_name]
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = old_score
    wb.save(wb_file)


def run_game(strategy, sample: int, sheet_name):
    global game
    global old_score
    browser = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    browser.get('https://play2048.co/')
    p = browser.current_window_handle
    time.sleep(1)
    browser.switch_to.window(p)
    for n in range(sample):
        try:
            game = browser.find_element(By.TAG_NAME, 'html')
        except NoSuchElementException:
            sys.exit("Unable to find the Game. Exiting program.")

        bad_char = ['+', '\n']
        old_score = 1
        left_try = 0
        while True:
            #added this only for Josh_strategy since it occasionally
            # got stuck and resulted in incorrect game over (I know, sloppy...)
            new_score = ''
            try:
                score = browser.find_element(
                    By.CLASS_NAME, 'score-container').text
                strategy()
                for char in score:
                    if char in bad_char:
                        break
                    else:
                        new_score += char
                if int(new_score) != old_score:
                    old_score = int(new_score)
                    continue
                else:
                    if left_try == 0:
                        game.send_keys(Keys.LEFT)
                        left_try += 1
                        continue
                    else:
                        update_sheet(sheet_name)
                        try:
                            restart = browser.find_element(
                                By.CLASS_NAME, 'restart-button')
                        except NoSuchElementException:
                            sys.exit("Unable to find the restart button. "
                                     "Exiting program.")
                        restart.click()
                        break
            except NoSuchElementException:
                sys.exit("Trouble reading score container. Exiting program")
    browser.quit()


folder = 'U:\\Joshua\\Dropbox\\Dropbox\\Python\\AutomatetheBoring\\'
old_score = 1
tries = get_integer('Input amount of tries:')
ad_block = f'{folder}Adblock-Plusfree-ad-blocker.crx'
options = Options()
options.add_extension(ad_block)
wb = Workbook()
wb_file = f'{folder}2048.xlsx'
run_game(josh_strategy, tries, 'Josh')
run_game(book_strategy, tries, 'Book')




# Code tested and works. Figuring out how to play the game was easy.
# Getting a game over status and score was not as simple.
# I found I could easily get the score itself though, so basically,
# I had this go through the Keys loop and
# then at the end compare new score with score of last loop,
# if scores didn't match. overwrite old score and continue.
# If scores match, break, and print final score and close program

# The score would sometimes output score + addition with new line.
# I just pulled out any + or \n char as we'd only get a game over
# after the game had stopped completely,
# looped one more time and checked that there was no change to score.
